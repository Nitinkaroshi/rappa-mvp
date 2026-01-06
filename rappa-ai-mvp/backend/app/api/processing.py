"""Processing status and results API endpoints."""

import logging
from typing import List, Optional
from io import BytesIO
from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse, Response
from sqlalchemy.orm import Session
from pydantic import BaseModel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.core.database import get_db
from app.core.auth import get_current_active_user
from app.models.user import User
from app.models.job import Job, JobStatus
from app.models.field import ExtractedField
from app.services.storage_service import StorageService

logger = logging.getLogger(__name__)

router = APIRouter()
storage_service = StorageService()


# ============================================================================
# Pydantic Schemas
# ============================================================================

class FieldResponse(BaseModel):
    """Response schema for extracted field."""
    field_name: str
    original_value: Optional[str]
    edited_value: Optional[str]
    confidence: Optional[str]
    is_edited: bool


class JobResponse(BaseModel):
    """Response schema for job details."""
    id: int
    filename: str
    file_type: str
    status: str
    error_message: Optional[str]
    created_at: str
    completed_at: Optional[str]


class JobResultsResponse(BaseModel):
    """Response schema for job results with extracted fields."""
    job: JobResponse
    fields: List[FieldResponse]
    fields_count: int


class JobListResponse(BaseModel):
    """Response schema for list of jobs."""
    jobs: List[JobResponse]
    total: int


# ============================================================================
# Processing Endpoints
# ============================================================================

@router.get("/status/{job_id}", response_model=JobResponse)
def get_job_status(
    job_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Get the status of a processing job.

    Args:
        job_id: ID of the job
        current_user: Authenticated user
        db: Database session

    Returns:
        JobResponse: Job status information

    Raises:
        HTTPException 404: If job not found
        HTTPException 403: If job belongs to another user
    """
    job = db.query(Job).filter(Job.id == job_id).first()

    if not job:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Job {job_id} not found"
        )

    # Check if job belongs to current user
    if job.user_id != current_user.id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You don't have permission to access this job"
        )

    return job.to_dict()


@router.get("/results/{job_id}", response_model=JobResultsResponse)
def get_job_results(
    job_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Get the results of a completed processing job.

    Args:
        job_id: ID of the job
        current_user: Authenticated user
        db: Database session

    Returns:
        JobResultsResponse: Job results with extracted fields

    Raises:
        HTTPException 404: If job not found
        HTTPException 403: If job belongs to another user
        HTTPException 400: If job is not completed
    """
    job = db.query(Job).filter(Job.id == job_id).first()

    if not job:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Job {job_id} not found"
        )

    # Check if job belongs to current user
    if job.user_id != current_user.id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You don't have permission to access this job"
        )

    # Check if job is completed
    if job.status != JobStatus.COMPLETED:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Job is not completed. Current status: {job.status.value}"
        )

    # Get extracted fields
    fields = db.query(ExtractedField).filter(ExtractedField.job_id == job_id).all()

    return JobResultsResponse(
        job=job.to_dict(),
        fields=[field.to_dict() for field in fields],
        fields_count=len(fields)
    )


@router.get("/results/{job_id}/enhanced")
def get_job_results_enhanced(
    job_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Get enhanced results with document type, key-value pairs, summary, and confidence.

    Args:
        job_id: ID of the job
        current_user: Authenticated user
        db: Database session

    Returns:
        dict: Enhanced results with structured data
    """
    job = db.query(Job).filter(Job.id == job_id).first()

    if not job:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"Job {job_id} not found")

    if job.user_id != current_user.id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Access denied")

    if job.status != JobStatus.COMPLETED:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Job not completed: {job.status.value}")

    # Get all extracted fields
    fields = db.query(ExtractedField).filter(ExtractedField.job_id == job_id).all()

    # Separate metadata fields (prefixed with _) from data fields
    metadata = {}
    extracted_data = []

    for field in fields:
        if field.field_name.startswith('_'):
            # Metadata field
            key = field.field_name[1:]  # Remove _ prefix
            metadata[key] = field.original_value
        else:
            # Regular data field
            extracted_data.append({
                "key": field.field_name,
                "value": field.original_value,
                "is_edited": field.is_edited,
                "confidence": field.confidence
            })

    return {
        "job": job.to_dict(),
        "document_type": metadata.get("document_type", "Unknown"),
        "confidence": float(metadata.get("confidence", 0.0)),
        "summary": metadata.get("summary", ""),
        "extracted_data": extracted_data,
        "metadata": {
            "method": metadata.get("method", "unknown"),
            "pdf_type": metadata.get("pdf_type", "unknown")
        }
    }


@router.get("/results/{job_id}/export/excel")
def export_results_to_excel(
    job_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Export extracted results to Excel file.

    Args:
        job_id: ID of the job
        current_user: Authenticated user
        db: Database session

    Returns:
        StreamingResponse: Excel file download
    """
    job = db.query(Job).filter(Job.id == job_id).first()

    if not job:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Job not found")

    if job.user_id != current_user.id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Access denied")

    if job.status != JobStatus.COMPLETED:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Job not completed")

    # Get all extracted fields
    fields = db.query(ExtractedField).filter(ExtractedField.job_id == job_id).all()

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Data"

    # Define header style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Add headers
    ws['A1'] = "Field Name"
    ws['B1'] = "Value"

    for cell in ['A1', 'B1']:
        ws[cell].fill = header_fill
        ws[cell].font = header_font
        ws[cell].alignment = header_alignment

    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50

    # Separate metadata and data
    metadata = {}
    extracted_data = []

    for field in fields:
        if field.field_name.startswith('_'):
            metadata[field.field_name[1:]] = field.original_value
        else:
            extracted_data.append((field.field_name, field.original_value))

    # Add metadata section
    row = 2
    ws[f'A{row}'] = "DOCUMENT INFORMATION"
    ws[f'A{row}'].font = Font(bold=True, size=11)
    row += 1

    ws[f'A{row}'] = "Document Type"
    ws[f'B{row}'] = metadata.get("document_type", "Unknown")
    row += 1

    ws[f'A{row}'] = "Confidence"
    ws[f'B{row}'] = f"{float(metadata.get('confidence', 0.0)):.1%}"
    row += 1

    ws[f'A{row}'] = "Summary"
    ws[f'B{row}'] = metadata.get("summary", "")
    ws[f'B{row}'].alignment = Alignment(wrap_text=True)
    row += 2

    # Add extracted data section
    ws[f'A{row}'] = "EXTRACTED FIELDS"
    ws[f'A{row}'].font = Font(bold=True, size=11)
    row += 1

    for field_name, value in extracted_data:
        ws[f'A{row}'] = field_name
        ws[f'B{row}'] = value
        row += 1

    # Save to bytes
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Return as downloadable file
    filename = f"{job.filename.rsplit('.', 1)[0]}_extracted_data.xlsx"

    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/jobs", response_model=JobListResponse)
def get_user_jobs(
    skip: int = 0,
    limit: int = 50,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Get all jobs for the current user.

    Args:
        skip: Number of records to skip (pagination)
        limit: Maximum number of records to return
        current_user: Authenticated user
        db: Database session

    Returns:
        JobListResponse: List of user's jobs
    """
    # Get total count
    total = db.query(Job).filter(Job.user_id == current_user.id).count()

    # Get jobs with pagination
    jobs = (
        db.query(Job)
        .filter(Job.user_id == current_user.id)
        .order_by(Job.created_at.desc())
        .offset(skip)
        .limit(limit)
        .all()
    )

    return JobListResponse(
        jobs=[job.to_dict() for job in jobs],
        total=total
    )


@router.delete("/{job_id}")
def delete_job(
    job_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Delete a job and its associated data.

    Note: This does NOT delete the file from S3 storage.

    Args:
        job_id: ID of the job to delete
        current_user: Authenticated user
        db: Database session

    Returns:
        dict: Deletion confirmation

    Raises:
        HTTPException 404: If job not found
        HTTPException 403: If job belongs to another user
    """
    job = db.query(Job).filter(Job.id == job_id).first()

    if not job:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Job {job_id} not found"
        )

    # Check if job belongs to current user
    if job.user_id != current_user.id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You don't have permission to delete this job"
        )

    # Delete job (cascades to extracted fields)
    db.delete(job)
    db.commit()

    logger.info(f"Job {job_id} deleted by user {current_user.id}")

    return {
        "message": "Job deleted successfully",
        "job_id": job_id
    }


@router.get("/stats")
def get_processing_stats(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Get processing statistics for the current user.

    Args:
        current_user: Authenticated user
        db: Database session

    Returns:
        dict: Processing statistics
    """
    # Count jobs by status
    total_jobs = db.query(Job).filter(Job.user_id == current_user.id).count()
    queued_jobs = db.query(Job).filter(
        Job.user_id == current_user.id,
        Job.status == JobStatus.QUEUED
    ).count()
    processing_jobs = db.query(Job).filter(
        Job.user_id == current_user.id,
        Job.status == JobStatus.PROCESSING
    ).count()
    completed_jobs = db.query(Job).filter(
        Job.user_id == current_user.id,
        Job.status == JobStatus.COMPLETED
    ).count()
    failed_jobs = db.query(Job).filter(
        Job.user_id == current_user.id,
        Job.status == JobStatus.FAILED
    ).count()

    return {
        "total_jobs": total_jobs,
        "queued": queued_jobs,
        "processing": processing_jobs,
        "completed": completed_jobs,
        "failed": failed_jobs,
        "credits_remaining": current_user.credits
    }


@router.get("/jobs/{job_id}/preview")
def get_job_document_preview(
    job_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Get the original document file for preview.

    Downloads the document from S3 and returns it for browser display.
    Supports PDF, JPG, PNG, TIFF formats.

    Args:
        job_id: ID of the job
        current_user: Authenticated user
        db: Database session

    Returns:
        Response: Document file with appropriate content type

    Raises:
        HTTPException 404: If job not found
        HTTPException 403: If job belongs to another user
    """
    job = db.query(Job).filter(Job.id == job_id).first()

    if not job:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Job {job_id} not found"
        )

    # Check if job belongs to current user
    if job.user_id != current_user.id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You don't have permission to access this job"
        )

    try:
        # Download file content from S3
        file_content = storage_service.get_file_content(job.s3_path)

        # Determine content type based on file extension
        file_ext = job.filename.lower().split('.')[-1]
        content_type_map = {
            'pdf': 'application/pdf',
            'jpg': 'image/jpeg',
            'jpeg': 'image/jpeg',
            'png': 'image/png',
            'tif': 'image/tiff',
            'tiff': 'image/tiff'
        }

        content_type = content_type_map.get(file_ext, 'application/octet-stream')

        # Return file with appropriate content type
        return Response(
            content=file_content,
            media_type=content_type,
            headers={
                "Content-Disposition": f"inline; filename={job.filename}",
                "Cache-Control": "public, max-age=3600"
            }
        )

    except Exception as e:
        logger.error(f"Error downloading document for job {job_id}: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to retrieve document: {str(e)}"
        )
