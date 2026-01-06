"""Export service for generating CSV, JSON, Excel, and PDF files from extracted fields."""

import csv
import json
import logging
from io import StringIO, BytesIO
from typing import List, Dict, Any
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from sqlalchemy.orm import Session

from app.models.job import Job
from app.models.field import ExtractedField
from app.models.custom_field import CustomField

logger = logging.getLogger(__name__)


class ExportService:
    """Service for exporting job results to various formats."""

    def __init__(self):
        """Initialize export service."""
        logger.info("ExportService initialized")

    def export_to_csv(self, job_id: int, db: Session) -> str:
        """Export job fields to CSV format.

        Args:
            job_id: Job ID
            db: Database session

        Returns:
            str: CSV content as string

        Raises:
            ValueError: If job not found
        """
        job = db.query(Job).filter(Job.id == job_id).first()
        if not job:
            raise ValueError(f"Job {job_id} not found")

        fields = db.query(ExtractedField).filter(
            ExtractedField.job_id == job_id
        ).all()

        # Get custom fields
        custom_fields = db.query(CustomField).filter(
            CustomField.job_id == job_id
        ).all()

        # Create CSV in memory
        output = StringIO()
        writer = csv.writer(output)

        # Header row
        writer.writerow([
            'Field Name',
            'Value',
            'Edited',
            'Original Value',
            'Field Type',
            'Source'
        ])

        # Data rows - Extracted fields
        for field in fields:
            # Skip metadata fields (starting with _)
            if field.field_name.startswith('_'):
                continue

            current_value = field.edited_value if field.is_edited else field.original_value
            is_edited = 'Yes' if field.is_edited else 'No'
            original = field.original_value if field.is_edited else ''

            writer.writerow([
                field.field_name,
                current_value or '',
                is_edited,
                original or '',
                'text',
                'Extracted'
            ])

        # Data rows - Custom fields
        for custom_field in custom_fields:
            writer.writerow([
                custom_field.field_name,
                custom_field.field_value or '',
                'No',
                '',
                custom_field.field_type,
                'Custom'
            ])

        csv_content = output.getvalue()
        logger.info(f"Exported {len(fields)} fields to CSV for job {job_id}")

        return csv_content

    def export_to_json(self, job_id: int, db: Session, include_metadata: bool = True) -> Dict[str, Any]:
        """Export job fields to JSON format.

        Args:
            job_id: Job ID
            db: Database session
            include_metadata: Include job metadata in export

        Returns:
            dict: JSON-serializable dictionary

        Raises:
            ValueError: If job not found
        """
        job = db.query(Job).filter(Job.id == job_id).first()
        if not job:
            raise ValueError(f"Job {job_id} not found")

        fields = db.query(ExtractedField).filter(
            ExtractedField.job_id == job_id
        ).all()

        # Get custom fields
        custom_fields = db.query(CustomField).filter(
            CustomField.job_id == job_id
        ).all()

        # Separate metadata and regular fields
        metadata = {}
        extracted_fields = {}
        custom_fields_data = {}

        for field in fields:
            if field.field_name.startswith('_'):
                # Metadata field (e.g., _document_type, _summary, _confidence)
                metadata[field.field_name[1:]] = field.original_value
            else:
                # Regular extracted field
                current_value = field.edited_value if field.is_edited else field.original_value
                extracted_fields[field.field_name] = {
                    "value": current_value or "",
                    "original_value": field.original_value or "",
                    "is_edited": field.is_edited
                }

        # Add custom fields
        for custom_field in custom_fields:
            custom_fields_data[custom_field.field_name] = {
                "value": custom_field.field_value or "",
                "field_type": custom_field.field_type,
                "created_at": custom_field.created_at.isoformat() if custom_field.created_at else None
            }

        # Build JSON structure
        export_data = {
            "job_id": job.id,
            "filename": job.filename,
            "template_id": getattr(job, 'template_id', None),  # Optional field
            "status": job.status.value if hasattr(job.status, 'value') else str(job.status),
            "processed_at": job.completed_at.isoformat() if job.completed_at else None,
            "fields": extracted_fields,
            "custom_fields": custom_fields_data
        }

        if include_metadata:
            export_data["metadata"] = metadata

        logger.info(f"Exported {len(extracted_fields)} fields to JSON for job {job_id}")

        return export_data

    def export_to_excel(self, job_id: int, db: Session) -> bytes:
        """Export job fields to Excel (XLSX) format.

        Args:
            job_id: Job ID
            db: Database session

        Returns:
            bytes: Excel file content

        Raises:
            ValueError: If job not found
        """
        job = db.query(Job).filter(Job.id == job_id).first()
        if not job:
            raise ValueError(f"Job {job_id} not found")

        fields = db.query(ExtractedField).filter(
            ExtractedField.job_id == job_id
        ).all()

        # Get custom fields
        custom_fields = db.query(CustomField).filter(
            CustomField.job_id == job_id
        ).all()

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Extracted Fields"

        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Write header
        headers = ['Field Name', 'Value', 'Edited', 'Original Value', 'Source']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Write extracted fields
        row = 2
        for field in fields:
            # Skip metadata fields
            if field.field_name.startswith('_'):
                continue

            current_value = field.edited_value if field.is_edited else field.original_value
            is_edited = 'Yes' if field.is_edited else 'No'
            original = field.original_value if field.is_edited else ''

            ws.cell(row=row, column=1, value=field.field_name)
            ws.cell(row=row, column=2, value=current_value or '')
            ws.cell(row=row, column=3, value=is_edited)
            ws.cell(row=row, column=4, value=original or '')
            ws.cell(row=row, column=5, value='Extracted')

            row += 1

        # Write custom fields
        for custom_field in custom_fields:
            ws.cell(row=row, column=1, value=custom_field.field_name)
            ws.cell(row=row, column=2, value=custom_field.field_value or '')
            ws.cell(row=row, column=3, value='No')
            ws.cell(row=row, column=4, value='')
            ws.cell(row=row, column=5, value='Custom')

            row += 1

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Max width 50
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        excel_content = output.getvalue()

        logger.info(f"Exported {row - 2} fields to Excel for job {job_id}")

        return excel_content

    def export_to_pdf(self, job_id: int, db: Session) -> bytes:
        """Export job fields to PDF format.

        Args:
            job_id: Job ID
            db: Database session

        Returns:
            bytes: PDF file content

        Raises:
            ValueError: If job not found
        """
        job = db.query(Job).filter(Job.id == job_id).first()
        if not job:
            raise ValueError(f"Job {job_id} not found")

        fields = db.query(ExtractedField).filter(
            ExtractedField.job_id == job_id
        ).all()

        # Get custom fields
        custom_fields = db.query(CustomField).filter(
            CustomField.job_id == job_id
        ).all()

        # Create PDF in memory
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter)
        elements = []

        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#4472C4'),
            spaceAfter=12,
            alignment=TA_CENTER
        )
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=12,
            textColor=colors.HexColor('#4472C4'),
            spaceAfter=6
        )

        # Title
        title = Paragraph(f"Document Extraction Report", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.2 * inch))

        # Job Info
        job_info_data = [
            ['Filename:', job.filename],
            ['Template:', job.template_id or 'Auto-detected'],
            ['Processed:', job.completed_at.strftime('%Y-%m-%d %H:%M:%S') if job.completed_at else 'N/A'],
            ['Status:', job.status.value if hasattr(job.status, 'value') else str(job.status)]
        ]

        job_info_table = Table(job_info_data, colWidths=[1.5*inch, 4.5*inch])
        job_info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E7E6E6')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
        ]))
        elements.append(job_info_table)
        elements.append(Spacer(1, 0.3 * inch))

        # Extracted Fields Header
        fields_heading = Paragraph("Extracted Fields", heading_style)
        elements.append(fields_heading)
        elements.append(Spacer(1, 0.1 * inch))

        # Filter out metadata fields
        regular_fields = [f for f in fields if not f.field_name.startswith('_')]

        if regular_fields or custom_fields:
            # Create table data
            table_data = [['Field Name', 'Value', 'Status', 'Source']]

            # Add extracted fields
            for field in regular_fields:
                current_value = field.edited_value if field.is_edited else field.original_value
                status = 'Edited' if field.is_edited else 'Original'

                table_data.append([
                    field.field_name,
                    current_value or '',
                    status,
                    'Extracted'
                ])

            # Add custom fields
            for custom_field in custom_fields:
                table_data.append([
                    custom_field.field_name,
                    custom_field.field_value or '',
                    'Custom',
                    'Custom'
                ])

            # Create and style table
            fields_table = Table(table_data, colWidths=[1.8*inch, 2.5*inch, 1*inch, 0.9*inch])
            fields_table.setStyle(TableStyle([
                # Header row
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                # Data rows
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('TOPPADDING', (0, 1), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                # Grid
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                # Alternating row colors
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
            ]))
            elements.append(fields_table)
        else:
            no_fields = Paragraph("No fields extracted", styles['Normal'])
            elements.append(no_fields)

        # Footer
        elements.append(Spacer(1, 0.5 * inch))
        footer_text = f"Generated by Rappa.AI on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        footer = Paragraph(footer_text, styles['Normal'])
        elements.append(footer)

        # Build PDF
        doc.build(elements)
        pdf_content = output.getvalue()

        logger.info(f"Exported {len(regular_fields)} fields to PDF for job {job_id}")

        return pdf_content

    def export_batch_to_csv(self, job_ids: List[int], db: Session) -> str:
        """Export multiple jobs to a single CSV file.

        Args:
            job_ids: List of job IDs
            db: Database session

        Returns:
            str: CSV content with all jobs

        Raises:
            ValueError: If any job not found
        """
        output = StringIO()
        writer = csv.writer(output)

        # Header row
        writer.writerow([
            'Job ID',
            'Filename',
            'Template',
            'Field Name',
            'Value',
            'Edited',
            'Original Value'
        ])

        for job_id in job_ids:
            job = db.query(Job).filter(Job.id == job_id).first()
            if not job:
                logger.warning(f"Job {job_id} not found, skipping")
                continue

            fields = db.query(ExtractedField).filter(
                ExtractedField.job_id == job_id
            ).all()

            for field in fields:
                if field.field_name.startswith('_'):
                    continue

                current_value = field.edited_value if field.is_edited else field.original_value
                is_edited = 'Yes' if field.is_edited else 'No'
                original = field.original_value if field.is_edited else ''

                writer.writerow([
                    job.id,
                    job.filename,
                    job.template_id or 'N/A',
                    field.field_name,
                    current_value or '',
                    is_edited,
                    original or ''
                ])

        csv_content = output.getvalue()
        logger.info(f"Batch exported {len(job_ids)} jobs to CSV")

        return csv_content

    def export_to_tally(self, job_id: int, db: Session) -> str:
        """Export job fields to Tally-compatible CSV format.

        Tally import format with columns:
        - Date (DD-MM-YYYY)
        - Voucher Type (Sales, Purchase, Payment, Receipt, Journal)
        - Voucher Number
        - Ledger Name (Party name)
        - Amount
        - Narration (Description)

        Args:
            job_id: Job ID
            db: Database session

        Returns:
            str: Tally-compatible CSV content as string

        Raises:
            ValueError: If job not found
        """
        job = db.query(Job).filter(Job.id == job_id).first()
        if not job:
            raise ValueError(f"Job {job_id} not found")

        fields = db.query(ExtractedField).filter(
            ExtractedField.job_id == job_id
        ).all()

        # Build field dictionary for easy lookup
        field_dict = {}
        for field in fields:
            if not field.field_name.startswith('_'):
                current_value = field.edited_value if field.is_edited else field.original_value
                field_dict[field.field_name.lower()] = current_value or ''

        # Create CSV in memory
        output = StringIO()
        writer = csv.writer(output)

        # Tally CSV Header
        writer.writerow([
            'Date',
            'Voucher Type',
            'Voucher Number',
            'Ledger Name',
            'Amount',
            'Narration'
        ])

        # Try to map extracted fields to Tally format
        # Common field name variations
        date_field = (
            field_dict.get('date') or
            field_dict.get('invoice date') or
            field_dict.get('bill date') or
            field_dict.get('transaction date') or
            datetime.now().strftime('%d-%m-%Y')
        )

        voucher_type = (
            field_dict.get('voucher type') or
            field_dict.get('transaction type') or
            'Sales'  # Default to Sales
        )

        voucher_number = (
            field_dict.get('invoice number') or
            field_dict.get('bill number') or
            field_dict.get('voucher number') or
            field_dict.get('receipt number') or
            f"DOC-{job_id}"
        )

        party_name = (
            field_dict.get('party name') or
            field_dict.get('customer name') or
            field_dict.get('vendor name') or
            field_dict.get('buyer name') or
            field_dict.get('seller name') or
            'Unknown Party'
        )

        amount = (
            field_dict.get('total amount') or
            field_dict.get('amount') or
            field_dict.get('grand total') or
            field_dict.get('net amount') or
            '0.00'
        )

        # Clean amount (remove currency symbols, commas)
        if isinstance(amount, str):
            amount = amount.replace('₹', '').replace('Rs', '').replace(',', '').strip()

        narration = (
            field_dict.get('narration') or
            field_dict.get('description') or
            field_dict.get('remarks') or
            f"Document processed from {job.filename}"
        )

        # Write data row
        writer.writerow([
            date_field,
            voucher_type,
            voucher_number,
            party_name,
            amount,
            narration
        ])

        csv_content = output.getvalue()
        logger.info(f"Exported job {job_id} to Tally CSV format")

        return csv_content

    def export_bank_statement_to_tally(self, job_id: int, db: Session) -> str:
        """Export bank statement transactions to Tally-compatible CSV format.

        This method handles bank statements with multiple transactions,
        properly categorizing debits as Payments and credits as Receipts.

        Args:
            job_id: Job ID containing bank statement data
            db: Database session

        Returns:
            str: Tally-compatible CSV with all transactions

        Raises:
            ValueError: If job not found
        """
        job = db.query(Job).filter(Job.id == job_id).first()
        if not job:
            raise ValueError(f"Job {job_id} not found")

        fields = db.query(ExtractedField).filter(
            ExtractedField.job_id == job_id
        ).all()

        # Build field dictionary
        field_dict = {}
        for field in fields:
            if not field.field_name.startswith('_'):
                current_value = field.edited_value if field.is_edited else field.original_value
                field_dict[field.field_name.lower()] = current_value or ''

        # Create CSV in memory
        output = StringIO()
        writer = csv.writer(output)

        # Tally CSV Header
        writer.writerow([
            'Date',
            'Voucher Type',
            'Voucher Number',
            'Ledger Name',
            'Amount',
            'Narration'
        ])

        # Get account details
        account_name = (
            field_dict.get('account holder name') or
            field_dict.get('account_holder_name') or
            'Bank Account'
        )

        account_number = field_dict.get('account number') or field_dict.get('account_number') or ''

        # Extract transaction details
        transaction_date = field_dict.get('transaction date') or field_dict.get('transaction_date')
        debit_amount = field_dict.get('debit amount') or field_dict.get('debit_amount')
        credit_amount = field_dict.get('credit amount') or field_dict.get('credit_amount')
        description = field_dict.get('transaction description') or field_dict.get('transaction_description') or field_dict.get('description')
        cheque_number = field_dict.get('cheque number') or field_dict.get('cheque_number') or field_dict.get('cheque/ref number')
        party_name = field_dict.get('party name') or field_dict.get('party_name')

        # If we have transaction data, create entry
        if transaction_date:
            # Determine voucher type based on debit/credit
            if debit_amount and debit_amount.strip() and debit_amount != '0' and debit_amount != '0.00':
                voucher_type = 'Payment'
                amount = debit_amount
            elif credit_amount and credit_amount.strip() and credit_amount != '0' and credit_amount != '0.00':
                voucher_type = 'Receipt'
                amount = credit_amount
            else:
                voucher_type = 'Journal'
                amount = '0.00'

            # Clean amount
            if isinstance(amount, str):
                amount = amount.replace('₹', '').replace('Rs', '').replace(',', '').strip()

            # Voucher number
            voucher_number = cheque_number or f"BANK-{job_id}"

            # Ledger name (party or account)
            ledger_name = party_name or account_name

            # Narration
            narration = description or f"Bank transaction from statement"
            if account_number:
                narration += f" - A/c: {account_number}"

            # Write transaction row
            writer.writerow([
                transaction_date,
                voucher_type,
                voucher_number,
                ledger_name,
                amount,
                narration
            ])

        csv_content = output.getvalue()
        logger.info(f"Exported bank statement job {job_id} to Tally CSV format")

        return csv_content


def get_export_service() -> ExportService:
    """Get ExportService instance.

    Returns:
        ExportService: Singleton export service instance
    """
    return ExportService()
